#coding:utf-8
from scipy import stats
import numpy as np
from pandas import Series,DataFrame
from openpyxl import load_workbook
import math
import uuid

class AnovaTestFile:
    def __init__(self, data_file):
        self.data_file = data_file
        self.wb = load_workbook(data_file)
        self.sheetnames = [s for s in self.wb.get_sheet_names()\
                           if 'result' not in s]
        self.sheets = [self.wb.get_sheet_by_name(s) for s in self.sheetnames]
        self.resultsheets = []

    def create_result_sheet(self):
        self.resultsheets = [self.wb.create_sheet(title='result_'+name)\
                             for name in self.sheetnames]

    def save(self,filepath):
        self.wb.save(filepath)

class AnovaTestSheet:
    def __init__(self,sheet):
        self.sheet = sheet
        self.sheetname = self.sheet.title
        self.max_row = self.sheet.max_row
        self.df = self._get_whole_dataframe()
        self.ptable = {}
        self.pheader = self._get_pheader()
        self.tags = {}

    @staticmethod
    def _product_random_value(mean,stdrange):
        b = np.random.uniform(*stdrange)
        a = b/math.sqrt(2)
        x1,x2 = mean-a, mean+a
        return x1,x2,b

    def _get_mean_value(self,row):
        if row % 2 == 1:
            return self.sheet['G'+str(row)].value
        else:
            return self.sheet['G'+str(row-1)].value

    def _set_od_value(self,row,x1,x2):
        if row % 2 == 1:
            self.sheet['F'+str(row)]=x1
            self.sheet['F'+str(row+1)]=x2

    def _get_stdev_value(self,row):
        if row % 2 == 1:
            return self.sheet['H'+str(row)].value
        else:
            return self.sheet['H'+str(row-1)].value

    def _set_stdev_value(self,row,stdev):
        if row % 2 == 1:
            self.sheet['H'+str(row)] = stdev

    def _get_one_row(self,row):
        time = self.sheet['A'+str(row)].value
        organ = self.sheet['B'+str(row)].value
        sp = self.sheet['C'+str(row)].value
        c = self.sheet['D'+str(row)].value
        rep = self.sheet['E'+str(row)].value
        od = self.sheet['F'+str(row)].value
        mean = self._get_mean_value(row)
        stdev = self._get_stdev_value(row)
        return Series([time,organ,sp,c,rep,od,mean,stdev],\
                      index=['time','organ','sp','c','rep','od','mean','stdev'])

    def _get_whole_dataframe(self):
        data={}
        for i in range(3,self.max_row+1):
            data[i]=self._get_one_row(i)
        return DataFrame(data).T

    def data_falsification(self,stdrange):
        for i in range(3,self.max_row+1,2):
            mean = self._get_mean_value(i)
            x1,x2,b=self._product_random_value(mean,stdrange)
            self._set_od_value(i,x1,x2)
            self._set_stdev_value(i,b)
        self.df = self._get_whole_dataframe()

    def _get_pheader(self):
        pheader = {}
        for i in range(3,self.max_row+1,2):
            gname=self.sheet['A'+str(i)].value+'_'+\
                      self.sheet['B'+str(i)].value+'_'+\
                      self.sheet['C'+str(i)].value+'_'+\
                      str(self.sheet['D'+str(i)].value)
            pheader[i] = gname
            pheader[i+1] = gname
        return pheader

    def produce_ptable(self):
        for (time,organ,sp),group_l1 in self.df.groupby(['time','organ','sp']):
            group_l2 = [g for c,g in group_l1.groupby(['c'])]
            i = group_l2[0].index[0]
            for m in range(5):
                g1 = group_l2[m]
                i1 = g1.index[0]
                tmp = {}
                for n in range(5):
                    g2 = group_l2[n]
                    i2 = g2.index[0]
                    f,p = stats.f_oneway(g1['od'],g2['od'])
                    tmp[i2]=p
                self.ptable[i1] = tmp
                self.ptable[i1+1] = tmp

    def add_tags(self):
        for i in range(3,self.max_row+1,10):
            g = {j:(self.df.ix[j])['mean'] for j in range(i,i+10,2)}
            g = sorted(g.items(), key=lambda d:d[1], reverse = True)
            gx = [j[0] for j in g]
            pt = [self.ptable[j[0]] for j in g]
            tmp = [[] for j in range(5)]
            tmp[0] = [g[j][0] for j in range(5)]
            for k in range(5):
                l = len(tmp[k])
                if l <= 1:
                    break
                i1 = gx.index(tmp[k][0])
                remove_list = []
                for m in range(1,l):
                    i2 = tmp[k][m]
                    p = pt[i1][i2]
                    if p <= 0.05:
                        remove_list.append(i2)
                        tmp[k+1].append(i2)
                for value in remove_list:
                    tmp[k].remove(value)
            tmp = [t for t in tmp if len(t)>0]
            if len(tmp)>1:
                for k in range(1,len(tmp)):
                    for m in range(len(tmp[k])):
                        i1 = gx.index(tmp[k][m])
                        for n in range(len(tmp[k-1])):
                            i2 = tmp[k-1][n]
                            p = pt[i1][i2]
                            if p>0.05 and i2 not in tmp[k]:
                                tmp[k].append(i2)
            marks = ['a','b','c','d','e']
            for k in range(len(tmp)):
                m = marks[k]
                for key in tmp[k]:
                    value = self.tags.get(key)
                    if value == None:
                        self.tags[key] = m
                    else:
                        self.tags[key] = value+m

    def write_to_sheet(self):
        for (i,tag) in self.tags.items():
            self.sheet['I'+str(i)] = tag
        for i in range(3,self.max_row+1,10):
            pheaders = [self.pheader[j] for j in range(i,i+10,2)]
            for k in range(5):
                self.sheet['J'+str(i+k+1)] = pheaders[k]
                self.sheet[chr(ord('K')+k)+str(i)]=pheaders[k]
            pvalues = [self.ptable[j] for j in range(i,i+10,2)]
            pvalues = [sorted(pv.items(),key=lambda d:d[0]) for pv in pvalues]
            for m in range(5):
                for n in range(m,5):
                    self.sheet[chr(ord('K')+m)+str(i+1+n)]=pvalues[m][n][1]

def main():
    data_file = input('请输入需要处理的文件名：')
    atf = AnovaTestFile(data_file)
    print('该文件一共有如下'+str(len(atf.sheetnames))+'张表格：')
    for name in atf.sheetnames:
        print(name)
    for sheet in atf.sheets:
        ats = AnovaTestSheet(sheet)
        r = input(ats.sheetname+'是否需要填充数据（y/n）：')
        if r == 'y':
            r = input(ats.sheetname+'->请输入标准差范围（以英文逗号隔开）：')
            x,y = r.split(',')
            x,y = float(x),float(y)
            ats.data_falsification((x,y))
        ats.produce_ptable()
        ats.add_tags()
        ats.write_to_sheet()
    result_file = data_file.split('.')[0]+'_result('\
                +str(uuid.uuid1())[:8]+').xlsx'
    atf.save(result_file)

def test():
    pass

if __name__ == "__main__":

    main()

