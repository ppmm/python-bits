#coding:utf-8
from scipy import stats
import numpy as np
from pandas import Series,DataFrame
from openpyxl import load_workbook
import math
import uuid
import os

def chart(data_ws,result_ws):
    pass

def _produc_random_value(mean,stdrange):
    b = np.random.uniform(*stdrange)
    a = b/math.sqrt(2)
    x1,x2 = mean-a, mean+a
    return x1,x2,b

def _set_od_value(ws,row,x1,x2):
    if row % 2 == 1:
        ws['F'+str(row)]=x1
        ws['F'+str(row+1)]=x2

def _get_mean_value(ws,row):
    if row % 2 == 1:
        return ws['G'+str(row)].value
    else:
        return ws['G'+str(row-1)].value

def _get_stdev_value(ws,row):
    if row % 2 == 1:
        return ws['H'+str(row)].value
    else:
        return ws['H'+str(row-1)].value

def _set_stdev_value(ws,row,stdev):
    if row % 2 == 1:
        ws['H'+str(row)] = stdev

def _get_one_row(ws,row):
    time = ws['A'+str(row)].value
    organ = ws['B'+str(row)].value
    sp = ws['C'+str(row)].value
    c = ws['D'+str(row)].value
    rep = ws['E'+str(row)].value
    od = ws['F'+str(row)].value
    mean = _get_mean_value(ws,row)
    stdev = _get_stdev_value(ws,row)
    return Series([time,organ,sp,c,rep,float(od),float(mean),stdev],\
                  index=['time','organ','sp','c','rep','od','mean','stdev'])

def get_whole_dataframe(ws):
    data={}
    for i in range(3,ws.max_row+1):
        data[i]=_get_one_row(ws,i)
    return DataFrame(data).T

def _fill_data_ws(ws,stdrange):
    for i in range(3,ws.max_row+1,2):
        mean = _get_mean_value(ws,i)
        x1,x2,b=_produc_random_value(mean,stdrange)
        _set_od_value(ws,i,x1,x2)
        _set_stdev_value(ws,i,b)

def _set_p_talbe_header(ws,result_ws):
    for i in range(3,ws.max_row+1,10):
        group = []
        for j in range(i,i+10,2):
            gname=ws['A'+str(j)].value+'_'+\
                      ws['B'+str(j)].value+'_'+\
                      ws['C'+str(j)].value+'_'+\
                      str(ws['D'+str(j)].value)
            group.append(gname)
        for k in range(5):
            result_ws['B'+str(i+k+1)]=group[k]
            result_ws[chr(ord('C')+k)+str(i)]=group[k]

    # for i in range(3,ws.max_row+1,20):
    #     group = []
    #     for j in range(i,i+10,2):
    #         gname=ws['A'+str(j)].value+'_'+\
    #                   ws['B'+str(j)].value+'_'+\
    #                   ws['C'+str(j)].value+'_'+\
    #                   ws['C'+str(j+10)].value+'_'+\
    #                   str(ws['D'+str(j)].value)
    #         group.append(gname)
    #     for k in range(5):
    #         result_ws['J'+str(i+2*k+6)] = group[k]

def produce_p_table(ws,result_ws):
    df = get_whole_dataframe(ws)
    _set_p_talbe_header(ws,result_ws)
    for (time,organ,sp),group_l1 in df.groupby(['time','organ','sp']):
        group_l2 = [g for c,g in group_l1.groupby(['c'])]
        i = group_l2[0].index[0]
        for m in range(5):
            for n in range(m+1,5):
                g1 = group_l2[m]
                g2 = group_l2[n]
                f,p = stats.f_oneway(g1['od'],g2['od'])
                result_ws[chr(ord('C')+m)+str(i+1+n)]=p

    # for (time,organ,c),group_l1 in df.groupby(['time','organ','c']):
    #     group_l2 = [g for c,g in group_l1.groupby(['sp'])]
    #     i = group_l2[0].index[0]
    #     g1 = group_l2[0]
    #     g2 = group_l2[1]
    #     f,p = stats.f_oneway(g1['od'],g2['od'])
    #     result_ws['K'+str(i+6)]=p

def calc(data_ws,result_ws):
        _fill_data_ws(data_ws,(0.1,0.6))

        for i in range(3,data_ws.max_row+1,10):
            group=[]
            for j in range(i,i+10,2):
                gname=data_ws['A'+str(j)].value+'_'+\
                      data_ws['B'+str(j)].value+'_'+\
                      data_ws['C'+str(j)].value+'_'+\
                      str(data_ws['D'+str(j)].value)
                group.append([gname,Series([data_ws['F'+str(j)].value,\
                                            data_ws['F'+str(j+1)].value])])
            for k in range(5):
                result_ws['B'+str(i+k+1)]=group[k][0]
                result_ws[chr(ord('C')+k)+str(i)]=group[k][0]
            for m in range(5):
                for n in range(m,5):
                    args = [group[m][1],group[n][1]]
                    f,p = stats.f_oneway(*args)
                    result_ws[chr(ord('C')+m)+str(i+1+n)]=p

def main():
    wb = load_workbook(filename = 'data/PODz.xlsx')
    salt = wb.get_sheet_by_name('salt')
    alkali = wb.get_sheet_by_name('alkali')
    salt_result = wb.create_sheet(title="salt_result")
    alkali_result = wb.create_sheet(title="alkali_result")
    calc(salt,salt_result)
    calc(alkali,alkali_result)
    wb.save(filename = 'data/PODz_result.xlsx')
    print('处理完成！')

def test(data_file,result_file):
    wb = load_workbook(data_file)
    sheetnames = wb.get_sheet_names()
    for name in sheetnames:
        sheet = wb.get_sheet_by_name(name)
        result_sheet = wb.create_sheet(title='result_'+name)
        r = input(name+'->请输入标准差范围（以英文逗号隔开）：')
        x,y = r.split(',')
        x,y = float(x),float(y)
        _fill_data_ws(sheet, (x,y))
        print(name+"->填充随机值完成！")
        produce_p_table(sheet, result_sheet)
        print(name+"->计算P值完成！")
    # salt = wb.get_sheet_by_name('salt')
    # alkali = wb.get_sheet_by_name('alkali')
    # salt_result = wb.create_sheet(title='salt_result')
    # alkali_result = wb.create_sheet(title="alkali_result")
    # _fill_data_ws(salt, stdrange)
    # _fill_data_ws(alkali, stdrange)
    # produce_p_table(salt, salt_result)
    # produce_p_table(alkali, alkali_result)
    wb.save(result_file)

def add_tags(result_file):
    wb = load_workbook(result_file)
    

if __name__ == "__main__":
    # main()
    data_file = 'data2/ggb (copy).xlsx'
    result_file = data_file.split('.')[0]+'_result('\
                  +str(uuid.uuid1())[:8]+').xlsx'
    test(data_file,result_file)
    print(data_file+':处理完成！')


